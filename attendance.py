import streamlit as st
import cv2
import numpy as np
import sqlite3
import os
import face_recognition
import pandas as pd
import smtplib
from email.message import EmailMessage
import time
from datetime import datetime
import schedule
import threading
import shutil
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
import speech_recognition as sr
from PIL import Image
import io

# Update credentials configuration
ADMIN_USERNAME = "Ratnakar"
ADMIN_PASSWORD = "1234"
GUARD_USERNAME = "guard"
GUARD_PASSWORD = "guard123"
DB_FILE = "attendance.db"

# Add this at the start of your main code
st.set_page_config(
    page_title="Labor Attendance System",
    page_icon="👥",
    layout="wide",
    initial_sidebar_state="expanded"
)

def init_db():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Add guards table
    cursor.execute('''CREATE TABLE IF NOT EXISTS guards (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      username TEXT UNIQUE,
                      password TEXT)''')
    
    # Insert default guard if not exists
    cursor.execute("INSERT OR IGNORE INTO guards (username, password) VALUES (?, ?)", 
                  (GUARD_USERNAME, GUARD_PASSWORD))
    
    # Existing attendance table
    cursor.execute('''CREATE TABLE IF NOT EXISTS attendance (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      name TEXT,
                      shift TEXT,
                      supervisor TEXT,
                      timestamp TEXT,
                      image BLOB)''')
    
    # New tables for supervisors and shifts
    cursor.execute('''CREATE TABLE IF NOT EXISTS supervisors (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      name TEXT UNIQUE,
                      email TEXT)''')
    
    cursor.execute('''CREATE TABLE IF NOT EXISTS shifts (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      name TEXT,
                      start_time TEXT,
                      end_time TEXT)''')
    
    # Add labor_info table
    cursor.execute('''CREATE TABLE IF NOT EXISTS labor_info (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      name TEXT UNIQUE,
                      contact TEXT,
                      id_type TEXT,
                      id_number TEXT,
                      emergency_contact TEXT,
                      supervisor TEXT,
                      id_proof_image BLOB)''')
    
    # Add frequent_laborers table
    cursor.execute('''CREATE TABLE IF NOT EXISTS frequent_laborers (
                      id INTEGER PRIMARY KEY AUTOINCREMENT,
                      name TEXT UNIQUE)''')
    
    # Insert default supervisor if none exists
    cursor.execute("INSERT OR IGNORE INTO supervisors (name, email) VALUES (?, ?)",
                  ("Default Supervisor", "supervisor@example.com"))
    
    # Insert default shift if none exists
    cursor.execute("INSERT OR IGNORE INTO shifts (name, start_time, end_time) VALUES (?, ?, ?)",
                  ("Morning Shift", "09:00", "17:00"))
    
    conn.commit()
    conn.close()

# Function to encode image
def encode_image(image):
    _, buffer = cv2.imencode(".jpg", image)
    return buffer.tobytes()

# Function to decode image
def decode_image(image_blob):
    nparr = np.frombuffer(image_blob, np.uint8)
    return cv2.imdecode(nparr, cv2.IMREAD_COLOR)

# Function to capture attendance
def capture_attendance(name, shift, supervisor, image):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    encoded_image = encode_image(image)
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO attendance (name, shift, supervisor, timestamp, image) VALUES (?, ?, ?, ?, ?)",
                   (name, shift, supervisor, timestamp, encoded_image))
    conn.commit()
    conn.close()

# Function to generate report
def generate_report():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Create a directory for temporary image storage
    report_dir = "report_images"
    os.makedirs(report_dir, exist_ok=True)
    
    # Create Excel workbook with better formatting
    wb = Workbook()
    
    # Attendance Sheet
    ws_attendance = wb.active
    ws_attendance.title = "Attendance Report"
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    headers = ["Name", "Shift", "Supervisor", "Timestamp", "Photo"]
    
    # Format headers for attendance sheet
    for col, header in enumerate(headers, 1):
        cell = ws_attendance.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_attendance.column_dimensions[get_column_letter(col)].width = 20
    
    # Add attendance data
    cursor.execute("""
        SELECT name, shift, supervisor, timestamp, image 
        FROM attendance 
        ORDER BY timestamp DESC
    """)
    rows = cursor.fetchall()
    
    for idx, (name, shift, supervisor, timestamp, image_blob) in enumerate(rows, 2):
        ws_attendance.cell(row=idx, column=1, value=name)
        ws_attendance.cell(row=idx, column=2, value=shift)
        ws_attendance.cell(row=idx, column=3, value=supervisor)
        ws_attendance.cell(row=idx, column=4, value=timestamp)
        
        if image_blob:
            image = decode_image(image_blob)
            image_path = f"{report_dir}/attendance_{idx}.jpg"
            cv2.imwrite(image_path, image)
            
            img = Image(image_path)
            img.width = 100
            img.height = 100
            ws_attendance.add_image(img, f'E{idx}')
            ws_attendance.row_dimensions[idx].height = 75
    
    # Labor Information Sheet
    ws_info = wb.create_sheet(title="Labor Information")
    
    # Headers for labor info sheet
    info_headers = ["Name", "Contact", "ID Type", "ID Number", "Emergency Contact", 
                   "Supervisor", "ID Proof"]
    
    for col, header in enumerate(info_headers, 1):
        cell = ws_info.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws_info.column_dimensions[get_column_letter(col)].width = 20
    
    # Add labor info data
    cursor.execute("SELECT * FROM labor_info")
    labor_info = cursor.fetchall()
    
    for idx, info in enumerate(labor_info, 2):
        for col, value in enumerate(info[1:7], 1):  # Skip id column
            ws_info.cell(row=idx, column=col, value=value)
        
        # Add ID proof image
        if info[7]:  # id_proof_image
            image = decode_image(info[7])
            image_path = f"{report_dir}/id_{idx}.jpg"
            cv2.imwrite(image_path, image)
            
            img = Image(image_path)
            img.width = 100
            img.height = 100
            ws_info.add_image(img, f'G{idx}')
            ws_info.row_dimensions[idx].height = 75
    
    # Save report
    report_file = "attendance_report.xlsx"
    wb.save(report_file)
    
    # Cleanup
    shutil.rmtree(report_dir)
    conn.close()
    
    return report_file

# Function to send email
def send_email(report_file):
    SMTP_SERVER = "smtp.gmail.com"
    SMTP_PORT = 587
    EMAIL_ADDRESS = "nikhilraval706@gmail.com"
    EMAIL_PASSWORD = "rhozifnjnmdgfvbt"
    TO_EMAIL = "22bcm045@nirmauni.ac.in"

    SUBJECT = "Labour Attendance Report"
    BODY = "Please find the attached attendance report."
    
    msg = EmailMessage()
    msg["From"] = EMAIL_ADDRESS
    msg["To"] = TO_EMAIL
    msg["Subject"] = SUBJECT
    msg.set_content(BODY)
    
    with open(report_file, "rb") as file:
        msg.add_attachment(file.read(), maintype="application", subtype="csv", filename=report_file)
    
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.send_message(msg)
        print("Email sent successfully!")
    except Exception as e:
        print(f"Error sending email: {e}")

# Schedule automated reports
def schedule_report():
    while True:
        schedule.run_pending()
        time.sleep(60)

schedule.every(8).hours.do(lambda: send_email(generate_report()))
threading.Thread(target=schedule_report, daemon=True).start()

def add_supervisor(name, email):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    try:
        cursor.execute("INSERT INTO supervisors (name, email) VALUES (?, ?)", (name, email))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def get_supervisors():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM supervisors")
    supervisors = [row[0] for row in cursor.fetchall()]
    conn.close()
    return supervisors

def add_shift(name, start_time, end_time):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO shifts (name, start_time, end_time) VALUES (?, ?, ?)",
                  (name, start_time, end_time))
    conn.commit()
    conn.close()

def get_shifts():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT name, start_time, end_time FROM shifts")
    shifts = cursor.fetchall()
    conn.close()
    return shifts

# Add function for managing frequent laborers
def add_frequent_laborer(name):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS frequent_laborers
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      name TEXT UNIQUE)''')
    try:
        cursor.execute("INSERT INTO frequent_laborers (name) VALUES (?)", (name,))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def get_frequent_laborers():
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM frequent_laborers")
    names = [row[0] for row in cursor.fetchall()]
    conn.close()
    return names

def delete_shift(shift_name):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM shifts WHERE name = ?", (shift_name,))
    conn.commit()
    conn.close()

def update_shift(shift_name, new_start_time, new_end_time):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE shifts 
        SET start_time = ?, end_time = ?
        WHERE name = ?
    """, (new_start_time, new_end_time, shift_name))
    conn.commit()
    conn.close()

# Add voice recognition function
def voice_to_text():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        st.write("Listening... Speak the labor's name")
        try:
            audio = r.listen(source, timeout=5)
            text = r.recognize_google(audio)
            return text
        except sr.UnknownValueError:
            st.error("Could not understand audio")
            return None
        except sr.RequestError:
            st.error("Could not request results")
            return None

# Add function to manage supervisors
def delete_supervisor(supervisor_name):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM supervisors WHERE name = ?", (supervisor_name,))
    conn.commit()
    conn.close()

def update_supervisor(old_name, new_name, new_email):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE supervisors 
        SET name = ?, email = ?
        WHERE name = ?
    """, (new_name, new_email, old_name))
    conn.commit()
    conn.close()

# Add new function for labor information
def add_labor_info(name, contact, id_type, id_number, emergency_contact, supervisor, image_blob):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Create labor_info table if not exists
    cursor.execute('''CREATE TABLE IF NOT EXISTS labor_info
                     (id INTEGER PRIMARY KEY AUTOINCREMENT,
                      name TEXT UNIQUE,
                      contact TEXT,
                      id_type TEXT,
                      id_number TEXT,
                      emergency_contact TEXT,
                      supervisor TEXT,
                      id_proof_image BLOB)''')
    
    try:
        cursor.execute("""
            INSERT INTO labor_info 
            (name, contact, id_type, id_number, emergency_contact, supervisor, id_proof_image)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (name, contact, id_type, id_number, emergency_contact, supervisor, image_blob))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()

def get_labor_info(name=None):
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    try:
        if name:
            cursor.execute("SELECT * FROM labor_info WHERE name = ?", (name,))
            info = cursor.fetchone()
        else:
            cursor.execute("SELECT * FROM labor_info")
            info = cursor.fetchall()
        return info if info else []
    except sqlite3.OperationalError:
        # If table doesn't exist, create it and return empty list
        cursor.execute('''CREATE TABLE IF NOT EXISTS labor_info (
                          id INTEGER PRIMARY KEY AUTOINCREMENT,
                          name TEXT UNIQUE,
                          contact TEXT,
                          id_type TEXT,
                          id_number TEXT,
                          emergency_contact TEXT,
                          supervisor TEXT,
                          id_proof_image BLOB)''')
        conn.commit()
        return []
    finally:
        conn.close()

def show_recent_entries():
    """
    Display recent attendance entries in a table format
    """
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Get today's date
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Fetch recent entries (last 10 entries from today)
    cursor.execute("""
        SELECT name, shift, supervisor, timestamp, image 
        FROM attendance 
        WHERE timestamp LIKE ? 
        ORDER BY timestamp DESC 
        LIMIT 10
    """, (f"{today}%",))
    
    entries = cursor.fetchall()
    conn.close()
    
    if entries:
        # Create a DataFrame for better display
        df = pd.DataFrame(
            [(name, shift, supervisor, timestamp.split()[1]) 
             for name, shift, supervisor, timestamp, _ in entries],
            columns=["Name", "Shift", "Supervisor", "Time"]
        )
        
        # Display the table with custom styling
        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
        )
        
        # Show attendance count
        st.metric("Total Attendance Today", len(entries))
        
        # Show entries with photos in expandable sections
        st.subheader("Detailed Entries")
        for name, shift, supervisor, timestamp, image_blob in entries:
            with st.expander(f"{name} - {timestamp}"):
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.write(f"**Shift:** {shift}")
                    st.write(f"**Supervisor:** {supervisor}")
                    st.write(f"**Time:** {timestamp}")
                with col2:
                    if image_blob:
                        image = decode_image(image_blob)
                        st.image(image, caption="Attendance Photo", width=150)
    else:
        st.info("No attendance records for today")

def show_dashboard():
    """
    Display admin dashboard with attendance statistics
    """
    st.subheader("Today's Overview")
    
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    
    # Get today's date
    today = datetime.now().strftime("%Y-%m-%d")
    
    # Get total attendance for today
    cursor.execute("""
        SELECT COUNT(*) 
        FROM attendance 
        WHERE timestamp LIKE ?
    """, (f"{today}%",))
    total_attendance = cursor.fetchone()[0]
    
    # Get shift-wise attendance
    cursor.execute("""
        SELECT shift, COUNT(*) 
        FROM attendance 
        WHERE timestamp LIKE ? 
        GROUP BY shift
    """, (f"{today}%",))
    shift_attendance = cursor.fetchall()
    
    # Get supervisor-wise attendance
    cursor.execute("""
        SELECT supervisor, COUNT(*) 
        FROM attendance 
        WHERE timestamp LIKE ? 
        GROUP BY supervisor
    """, (f"{today}%",))
    supervisor_attendance = cursor.fetchall()
    
    conn.close()
    
    # Display metrics
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Attendance", total_attendance)
    with col2:
        if shift_attendance:
            st.subheader("Shift-wise Count")
            for shift, count in shift_attendance:
                st.metric(shift, count)
    with col3:
        if supervisor_attendance:
            st.subheader("Supervisor-wise Count")
            for supervisor, count in supervisor_attendance:
                st.metric(supervisor, count)
    
    # Show recent entries
    st.subheader("Recent Entries")
    show_recent_entries()

# Modified Streamlit UI
def main():
    st.title("Labour Attendance System")
    
    # Initialize database
    init_db()
    
    # Initialize session state
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
    if 'user_role' not in st.session_state:
        st.session_state.user_role = None
    if 'form_data' not in st.session_state:
        st.session_state.form_data = {
            'name': '',
            'shift': '',
            'supervisor': ''
        }

    # Login Section
    if not st.session_state.logged_in:
        st.subheader("Login")
        login_type = st.radio("Select Login Type", ["Admin", "Guard"])
        
        with st.form("login_form"):
            username = st.text_input("Username")
            password = st.text_input("Password", type="password")
            submit = st.form_submit_button("Login")
            
            if submit:
                if login_type == "Admin" and username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
                    st.session_state.logged_in = True
                    st.session_state.user_role = "admin"
                    st.success("Admin login successful!")
                    st.rerun()
                elif login_type == "Guard" and username == GUARD_USERNAME and password == GUARD_PASSWORD:
                    st.session_state.logged_in = True
                    st.session_state.user_role = "guard"
                    st.success("Guard login successful!")
                    st.rerun()
                else:
                    st.error("Invalid credentials")
    
    else:
        # Admin Interface
        if st.session_state.user_role == "admin":
            admin_interface()
        
        # Guard Interface
        elif st.session_state.user_role == "guard":
            guard_interface()

def admin_interface():
    tabs = st.tabs(["Dashboard", "Labor Information", "Manage Supervisors", "Manage Shifts", "Reports"])
    
    with tabs[0]:
        show_dashboard()
    
    with tabs[1]:
        st.header("Labor Information Management")
        
        with st.form("add_labor_info"):
            name = st.text_input("Labor Name")
            contact = st.text_input("Contact Number")
            id_type = st.selectbox("ID Type", ["Aadhar Card", "PAN Card", "Voter ID", "Other"])
            id_number = st.text_input("ID Number")
            emergency_contact = st.text_input("Emergency Contact")
            supervisor = st.selectbox("Point of Contact Supervisor", get_supervisors())
            
            # Photo capture option
            photo_method = st.radio("Choose ID Proof Method", ["Upload Photo", "Capture Photo"])
            
            if photo_method == "Upload Photo":
                id_proof = st.file_uploader("Upload ID Proof", type=['jpg', 'jpeg', 'png'])
                image_bytes = id_proof.getvalue() if id_proof else None
            else:
                id_proof = st.camera_input("Capture ID Proof")
                image_bytes = id_proof.getvalue() if id_proof else None
            
            if st.form_submit_button("Add Labor Information"):
                if image_bytes:
                    if add_labor_info(name, contact, id_type, id_number, 
                                    emergency_contact, supervisor, image_bytes):
                        st.success("Labor information added successfully!")
                    else:
                        st.error("Labor already exists in records!")
                else:
                    st.error("Please provide ID proof photo!")
        
        st.subheader("Existing Labor Records")
        labor_records = get_labor_info()
        if labor_records:
            for record in labor_records:
                with st.expander(f"Labor: {record[1]}"):
                    st.write(f"Contact: {record[2]}")
                    st.write(f"ID Type: {record[3]}")
                    st.write(f"ID Number: {record[4]}")
                    st.write(f"Emergency Contact: {record[5]}")
                    st.write(f"Supervisor: {record[6]}")
                    if record[7]:  # ID proof image
                        st.image(decode_image(record[7]), caption="ID Proof")
    
    with tabs[2]:
        st.header("Manage Supervisors")
        with st.form("add_supervisor"):
            sup_name = st.text_input("Supervisor Name")
            sup_email = st.text_input("Supervisor Email")
            if st.form_submit_button("Add Supervisor"):
                if add_supervisor(sup_name, sup_email):
                    st.success("Supervisor added successfully!")
                else:
                    st.error("Supervisor already exists!")
        
        st.subheader("Current Supervisors")
        st.write(pd.DataFrame({"Supervisors": get_supervisors()}))
    
    with tabs[3]:
        st.header("Manage Shifts")
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("Add New Shift")
            with st.form("add_shift"):
                shift_name = st.text_input("Shift Name")
                start_time = st.time_input("Start Time", key="new_shift_start")
                end_time = st.time_input("End Time", key="new_shift_end")
                if st.form_submit_button("Add Shift"):
                    add_shift(shift_name, start_time.strftime("%H:%M"), end_time.strftime("%H:%M"))
                    st.success("Shift added successfully!")
        
        with col2:
            st.subheader("Edit/Delete Shifts")
            shifts = get_shifts()
            for idx, shift in enumerate(shifts):
                with st.expander(f"Shift: {shift[0]}"):
                    new_start = st.time_input(
                        f"Start Time for {shift[0]}", 
                        value=datetime.strptime(shift[1], "%H:%M").time(),
                        key=f"start_{idx}"
                    )
                    new_end = st.time_input(
                        f"End Time for {shift[0]}", 
                        value=datetime.strptime(shift[2], "%H:%M").time(),
                        key=f"end_{idx}"
                    )
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button(f"Update {shift[0]}", key=f"update_{idx}"):
                            update_shift(shift[0], new_start.strftime("%H:%M"), 
                                      new_end.strftime("%H:%M"))
                            st.success("Shift updated successfully!")
                    with col2:
                        if st.button(f"Delete {shift[0]}", key=f"delete_{idx}"):
                            delete_shift(shift[0])
                            st.success("Shift deleted successfully!")
                            st.rerun()
    
    with tabs[4]:
        st.header("Generate Report")
        if st.button("Generate and Download Report"):
            report_file = generate_report()
            send_email(report_file)
            
            # Download report
            with open(report_file, "rb") as file:
                st.download_button(
                    label="Download Report",
                    data=file,
                    file_name="attendance_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            st.success("Report generated and sent via email!")

def guard_interface():
    st.header("Attendance Capture")
    
    # Initialize session state for form data
    if 'form_data' not in st.session_state:
        st.session_state.form_data = {
            'name': '',
            'shift': '',
            'supervisor': ''
        }
    
    col1, col2 = st.columns(2)
    
    with col1:
        name = st.text_input("Labour Name", value=st.session_state.form_data.get('name', ''))
        # Add voice input button
        if st.button("🎤", key="voice_input"):
            name = voice_to_text()
            if name:
                st.session_state.form_data['name'] = name
                st.rerun()
        
        # Update session state with current name
        st.session_state.form_data['name'] = name
        
        shift = st.selectbox("Select Shift", 
                           [shift[0] for shift in get_shifts()],
                           key="shift_select")
        st.session_state.form_data['shift'] = shift
        
        supervisor = st.selectbox("Select Supervisor", 
                                get_supervisors(),
                                key="supervisor_select")
        st.session_state.form_data['supervisor'] = supervisor
        
        # Real-time photo capture only
        st.write("Capture Photo (Required)")
        picture = st.camera_input("Take a picture")
        
        # Form validation
        form_complete = (name and shift and supervisor and picture is not None)
        
        if not form_complete:
            st.warning("Please fill all fields and capture photo")
        else:
            if st.button("Submit Attendance"):
                image = cv2.imdecode(np.frombuffer(picture.getvalue(), np.uint8), 
                                   cv2.IMREAD_COLOR)
                # Face detection
                face_locations = face_recognition.face_locations(image)
                if face_locations:
                    capture_attendance(name, shift, supervisor, image)
                    st.success("Attendance recorded successfully!")
                    # Add to frequent laborers if not exists
                    add_frequent_laborer(name)
                    # Clear form data after successful submission
                    st.session_state.form_data = {
                        'name': '',
                        'shift': '',
                        'supervisor': ''
                    }
                    st.rerun()
                else:
                    st.error("No face detected in the photo!")
    
    with col2:
        st.subheader("Recent Entries")
        show_recent_entries()

if __name__ == "__main__":
    init_db()
    main()
